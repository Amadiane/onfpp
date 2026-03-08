
# views.py — fichier COMPLET

from rest_framework import generics, viewsets, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView

from .models import User, Role, Region, Centre
from .serializers import (
    UserCreateSerializer, CustomTokenObtainPairSerializer,
    RoleSerializer, RegionSerializer, CentreSerializer,
)


# ── Login JWT ────────────────────────────────────────────────
# class CustomLoginView(TokenObtainPairView):
#     serializer_class = CustomTokenObtainPairSerializer

#     def post(self, request, *args, **kwargs):
#         print("LOGIN DATA:", request.data)
#         return super().post(request, *args, **kwargs)


# # ── Utilisateur connecté ─────────────────────────────────────
# class MeView(APIView):
#     permission_classes = [IsAuthenticated]

#     def get(self, request):
#         user = request.user
#         return Response({
#             "id":         user.id,
#             "username":   user.username,
#             "first_name": user.first_name,
#             "last_name":  user.last_name,
#             "email":      user.email,
#             "role":       user.role.name   if user.role   else None,
#             "niveau":     user.role.level  if user.role   else 0,
#             "region":     user.region.name if user.region else None,
#             "centre":     user.centre.name if user.centre else None,
#         })

# views_auth.py  — À ajouter dans ton fichier views.py principal
#
# Problème : /api/me/ retourne 403 car la vue n'est pas protégée
# correctement OU le token n'est pas lu par Django.
# Solution : utiliser IsAuthenticated + SimpleJWT correctement.

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from rest_framework import status
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import authenticate
from django.contrib.auth import get_user_model

User = get_user_model()


# ── 1. LOGIN ────────────────────────────────────────────────────────
# POST /api/login/
# Retourne access + refresh + infos utilisateur complètes

@api_view(["POST"])
@permission_classes([AllowAny])
def login_view(request):
    username = request.data.get("username", "").strip()
    password = request.data.get("password", "")

    if not username or not password:
        return Response(
            {"detail": "Nom d'utilisateur et mot de passe requis."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    user = authenticate(request, username=username, password=password)

    if user is None:
        return Response(
            {"detail": "Identifiants incorrects."},
            status=status.HTTP_401_UNAUTHORIZED,
        )

    if not user.is_active:
        return Response(
            {"detail": "Ce compte est désactivé."},
            status=status.HTTP_403_FORBIDDEN,
        )

    # Génère les tokens JWT
    refresh = RefreshToken.for_user(user)
    access  = str(refresh.access_token)

    return Response({
        "access":   access,
        "refresh":  str(refresh),
        "username": user.username,
        "role":     getattr(user, "role",   None),
        "niveau":   getattr(user, "niveau", 0),
        "region":   getattr(user, "region", None),
        "centre":   getattr(user, "centre", None),
    })


# ── 2. ME ────────────────────────────────────────────────────────────
# GET /api/me/
# Retourne le profil complet de l'utilisateur connecté.
# ⚠️  NÉCESSITE : permission_classes = [IsAuthenticated]
#     + DEFAULT_AUTHENTICATION_CLASSES = JWTAuthentication dans settings.py

@api_view(["GET"])
@permission_classes([IsAuthenticated])   # ← NE PAS mettre AllowAny ici
def me_view(request):
    user = request.user
    return Response({
        "id":         user.id,
        "username":   user.username,
        "first_name": user.first_name,
        "last_name":  user.last_name,
        "email":      user.email,
        "role":       getattr(user, "role",   None),
        "niveau":     getattr(user, "niveau", 0),
        "region":     getattr(user, "region", None),
        "centre":     getattr(user, "centre", None),
        "is_staff":   user.is_staff,
    })


# ── Créer un utilisateur ─────────────────────────────────────
class UserCreateView(generics.CreateAPIView):
    serializer_class   = UserCreateSerializer
    permission_classes = [AllowAny]   # ← remplacez par IsAuthenticated en prod


# ── Roles / Régions / Centres (lecture seule) ────────────────
class RoleViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = Role.objects.all()
    serializer_class   = RoleSerializer
    permission_classes = [IsAuthenticated]


class RegionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = Region.objects.all()
    serializer_class   = RegionSerializer
    permission_classes = [IsAuthenticated]


class CentreViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = Centre.objects.all()
    serializer_class   = CentreSerializer
    permission_classes = [IsAuthenticated]






# views.py — fichier COMPLET

from rest_framework import generics, viewsets, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework_simplejwt.views import TokenObtainPairView

from .models import User, Role, Region, Centre
from .serializers import (
    UserCreateSerializer, CustomTokenObtainPairSerializer,
    RoleSerializer, RegionSerializer, CentreSerializer,
)


# ── Login JWT ────────────────────────────────────────────────
class CustomLoginView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer


# ── Utilisateur connecté ─────────────────────────────────────
class MeView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        return Response({
            "id":         user.id,
            "username":   user.username,
            "first_name": user.first_name,
            "last_name":  user.last_name,
            "email":      user.email,
            "role":       user.role.name   if user.role   else None,
            "niveau":     user.role.level  if user.role   else 0,
            "region":     user.region.name if user.region else None,
            "centre":     user.centre.name if user.centre else None,
        })


# ── Créer un utilisateur ─────────────────────────────────────
class UserCreateView(generics.CreateAPIView):
    serializer_class   = UserCreateSerializer
    permission_classes = [AllowAny]   # ← remplacez par IsAuthenticated en prod


# ── Roles / Régions / Centres (lecture seule) ────────────────
class RoleViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = Role.objects.all()
    serializer_class   = RoleSerializer
    # permission_classes = [IsAuthenticated]
    permission_classes = [AllowAny]  


class RegionViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = Region.objects.all()
    serializer_class   = RegionSerializer
    permission_classes = [IsAuthenticated]


class CentreViewSet(viewsets.ReadOnlyModelViewSet):
    queryset           = Centre.objects.all()
    serializer_class   = CentreSerializer
    permission_classes = [IsAuthenticated]


# ── Liste des utilisateurs (GET /api/users/) ─────────────────
from rest_framework import serializers as drf_serializers

class UserListSerializer(drf_serializers.ModelSerializer):
    role   = RoleSerializer(read_only=True)
    region = RegionSerializer(read_only=True)
    centre = CentreSerializer(read_only=True)

    class Meta:
        model  = User
        fields = [
            "id", "username", "first_name", "last_name",
            "email", "is_active", "date_joined",
            "role", "region", "centre",
        ]

class UserListView(generics.ListAPIView):
    queryset           = User.objects.select_related("role","region","centre").all().order_by("-date_joined")
    serializer_class   = UserListSerializer
    permission_classes = [IsAuthenticated]


# ── Liste des utilisateurs ───────────────────────────────────
from rest_framework import serializers as drf_serializers

class UserListSerializer(drf_serializers.ModelSerializer):
    role   = RoleSerializer(read_only=True)
    region = RegionSerializer(read_only=True)
    centre = CentreSerializer(read_only=True)

    class Meta:
        model  = User
        fields = [
            "id", "username", "first_name", "last_name",
            "email", "is_active", "role", "region", "centre",
        ]

class UserListView(generics.ListAPIView):
    queryset           = User.objects.select_related("role","region","centre").all()
    serializer_class   = UserListSerializer
    # permission_classes = [IsAuthenticated]
    permission_classes = [AllowAny]  




























# views.py
# views.py — VERSION PROPRE (zéro matplotlib/numpy/seaborn)
# views.py — VERSION PROPRE (zéro matplotlib/numpy/seaborn)

# views.py — VERSION PROPRE (zéro matplotlib/numpy/seaborn)
# views.py — VERSION PROPRE (zéro matplotlib/numpy/seaborn)
# views.py — VERSION PROPRE (zéro matplotlib/numpy/seaborn)
# views.py — VERSION PROPRE (zéro matplotlib/numpy/seaborn)

import io

from django.http import HttpResponse
from rest_framework.viewsets import ModelViewSet
from rest_framework.decorators import action
from rest_framework.response import Response

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import EvaluationSession, Apprenant, Critere, Evaluation, NOTE_MAPPING
from .serializers import (
    EvaluationSessionSerializer,
    CritereSerializer,
    ApprenantSerializer,
    EvaluationSerializer,
)

# ─── Couleurs ONFPP ───────────────────────────────────────
NAVY   = colors.HexColor("#0D1B5E")
BLUE   = colors.HexColor("#1A3BD4")
ICE    = colors.HexColor("#C8D9FF")
GREEN  = colors.HexColor("#0BA376")
ORANGE = colors.HexColor("#F5A800")
RED    = colors.HexColor("#E53935")
WHITE  = colors.white
LIGHT  = colors.HexColor("#EEF2FF")

def score_color(pct):
    if pct >= 75: return GREEN
    if pct >= 50: return ORANGE
    return RED

def score_label(pct):
    if pct >= 75: return "Très satisfaisant"
    if pct >= 50: return "Satisfaisant"
    return "Insuffisant"


# ─── CRUD simples ─────────────────────────────────────────
class CritereViewSet(ModelViewSet):
    queryset = Critere.objects.all()
    serializer_class = CritereSerializer

class ApprenantViewSet(ModelViewSet):
    queryset = Apprenant.objects.all()
    serializer_class = ApprenantSerializer

class EvaluationViewSet(ModelViewSet):
    queryset = Evaluation.objects.all()
    serializer_class = EvaluationSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        sid = self.request.query_params.get("session")
        if sid:
            qs = qs.filter(session_id=sid)
        return qs

    @action(detail=True, methods=["patch"], url_path="update-commentaire")
    def update_commentaire(self, request, pk=None):
        ev = self.get_object()
        commentaire = request.data.get("commentaire", "")
        ev.commentaire = commentaire
        ev.save(update_fields=["commentaire"])
        return Response({"id": ev.id, "commentaire": ev.commentaire})


# ─── Session ──────────────────────────────────────────────
class EvaluationSessionViewSet(ModelViewSet):
    queryset = EvaluationSession.objects.all()
    serializer_class = EvaluationSessionSerializer

    @action(detail=True, methods=["patch"], url_path="update-commentaire-final")
    def update_commentaire_final(self, request, pk=None):
        session = self.get_object()
        commentaire_final = request.data.get("commentaire_final", "")
        session.commentaire_final = commentaire_final
        session.save(update_fields=["commentaire_final"])
        return Response({"id": session.id, "commentaire_final": session.commentaire_final})

    # ── JSON résultats ──────────────────────────────────
    @action(detail=True, methods=["get"])
    def results(self, request, pk=None):
        session = self.get_object()
        apprenants = Apprenant.objects.filter(
            id__in=session.evaluation_set.values_list("apprenant", flat=True).distinct()
        )
        data = []
        for a in apprenants:
            total   = session.total_points_apprenant(a)
            percent = session.percentage_apprenant(a)
            data.append({
                "apprenant_id": a.id,
                "apprenant":    a.nom,
                "total_points": total,
                "percentage":   percent,
            })
        return Response(data)

    # ── JSON pour graphiques frontend ───────────────────
    @action(detail=True, methods=["get"], url_path="graph")
    def graph(self, request, pk=None):
        session = self.get_object()
        apprenants = Apprenant.objects.filter(
            id__in=session.evaluation_set.values_list("apprenant", flat=True).distinct()
        )
        criteres = Critere.objects.all()

        # Scores par apprenant
        scores = []
        for a in apprenants:
            total   = session.total_points_apprenant(a)
            percent = session.percentage_apprenant(a)
            scores.append({"apprenant": a.nom, "pourcentage": percent, "total": total})

        # Comptage notes globales (pour camembert satisfaction)
        notes_counter = {1: 0, 2: 0, 3: 0}
        for ev in session.evaluation_set.all():
            if ev.note in notes_counter:
                notes_counter[ev.note] += 1

        # Scores par critère
        critere_data = []
        nb_app = apprenants.count()
        for c in criteres:
            evs = session.evaluation_set.filter(critere=c)
            pts = sum(NOTE_MAPPING.get(e.note, 0) for e in evs)
            pct = round(pts / (nb_app * 75) * 100, 1) if nb_app else 0
            critere_data.append({"nom": c.nom, "total": pts, "pourcentage": pct})

        return Response({
            "scores":          scores,
            "notes_counter":   notes_counter,
            "critere_data":    critere_data,
        })

    # ── Export Excel ────────────────────────────────────
    @action(detail=True, methods=["get"])
    def export_excel(self, request, pk=None):
        session   = self.get_object()
        apprenants = Apprenant.objects.filter(
            id__in=session.evaluation_set.values_list("apprenant", flat=True).distinct()
        )
        criteres  = Critere.objects.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Résultats"

        hdr_font = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill("solid", fgColor="0D1B5E")
        c_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)

        headers = (["#", "Apprenant"] +
                   [c.nom for c in criteres] +
                   ["Total pts", "Max pts", "%", "Appréciation"])
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = c_align
            ws.column_dimensions[get_column_letter(col)].width = max(12, len(h) + 2)

        nb_crit = criteres.count()
        max_pts = nb_crit * 75
        ri = 2
        for i, a in enumerate(apprenants, 1):
            evals   = session.evaluation_set.filter(apprenant=a)
            total   = session.total_points_apprenant(a)
            percent = session.percentage_apprenant(a)
            appreci = score_label(percent)

            alt = PatternFill("solid", fgColor="EEF2FF" if i % 2 == 0 else "FFFFFF")
            row_data = [i, a.nom]
            for c in criteres:
                ev = evals.filter(critere=c).first()
                row_data.append(NOTE_MAPPING.get(ev.note, 0) if ev else 0)
            row_data += [total, max_pts, f"{percent}%", appreci]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=col, value=val)
                cell.fill = alt; cell.alignment = c_align
            ri += 1

        # Ligne totaux
        nb_app      = apprenants.count()
        total_glob  = sum(session.total_points_apprenant(a) for a in apprenants)
        max_glob    = nb_crit * 75 * nb_app
        pct_glob    = round(total_glob / max_glob * 100, 2) if max_glob else 0
        tc = len(headers) - 3
        ws.cell(row=ri, column=2, value="TOTAL SESSION")
        ws.cell(row=ri, column=tc,   value=total_glob)
        ws.cell(row=ri, column=tc+1, value=max_glob)
        ws.cell(row=ri, column=tc+2, value=f"{pct_glob}%")
        for col in range(1, len(headers)+1):
            cell = ws.cell(row=ri, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1A3BD4")
            cell.alignment = c_align

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="resultats_session_{session.id}.xlsx"'
        )
        wb.save(response)
        return response

    # ── PDF Global (sans graphiques, propre) ────────────
    @action(detail=True, methods=["get", "post"], url_path="pdf-global")
    def pdf_global(self, request, pk=None):
        session   = self.get_object()
        apprenants = Apprenant.objects.filter(
            id__in=session.evaluation_set.values_list("apprenant", flat=True).distinct()
        )
        criteres  = Critere.objects.all()
        nb_crit   = criteres.count()

        # Calculs
        data_app = []
        total_glob = 0
        for a in apprenants:
            total   = session.total_points_apprenant(a)
            percent = session.percentage_apprenant(a)
            total_glob += total
            data_app.append({"nom": a.nom, "total": total, "percent": percent})

        nb_app   = apprenants.count()
        max_glob = nb_crit * 75 * nb_app
        pct_glob = round(total_glob / max_glob * 100, 2) if max_glob else 0

        # PDF
        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            leftMargin=1.8*cm, rightMargin=1.8*cm,
            topMargin=2*cm,    bottomMargin=2*cm,
        )
        st = getSampleStyleSheet()
        el = []

        s_title = ParagraphStyle("T", parent=st["Title"],
                                 textColor=NAVY, fontSize=18,
                                 alignment=TA_CENTER, spaceAfter=4)
        s_sub   = ParagraphStyle("S", parent=st["Normal"],
                                 textColor=BLUE, fontSize=12,
                                 alignment=TA_CENTER, spaceAfter=10)
        s_h2    = ParagraphStyle("H2", parent=st["Heading2"],
                                 textColor=NAVY, fontSize=12,
                                 spaceBefore=14, spaceAfter=6)

        fd = lambda d: d.strftime("%d/%m/%Y") if d else "—"

        # En-tête
        el.append(Paragraph("RAPPORT D'ÉVALUATION DE FORMATION", s_title))
        el.append(Paragraph(session.theme, s_sub))
        el.append(Paragraph(
            f"Formateur : <b>{session.formateur}</b> &nbsp;·&nbsp; "
            f"Lieu : <b>{session.lieu}</b> &nbsp;·&nbsp; "
            f"Période : <b>{fd(session.periode_debut)} → {fd(session.periode_fin)}</b>",
            ParagraphStyle("meta", parent=st["Normal"],
                           alignment=TA_CENTER, fontSize=9,
                           textColor=NAVY, spaceAfter=10)
        ))
        el.append(HRFlowable(width="100%", thickness=3, color=NAVY))
        el.append(Spacer(1, 0.4*cm))

        # Résumé global
        appreci_g = score_label(pct_glob)
        sc_g = score_color(pct_glob)
        rt = Table([
            ["Apprenants évalués", str(nb_app),
             "Total points", str(total_glob)],
            ["Rubriques évaluées", str(nb_crit),
             "Points maximum",     str(max_glob)],
            ["Pourcentage global", f"{pct_glob} %",
             "Appréciation",       appreci_g],
        ], colWidths=[4.5*cm, 3*cm, 4.5*cm, 4.5*cm])
        rt.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (0,-1), ICE),
            ("BACKGROUND",   (2,0), (2,-1), ICE),
            ("FONTNAME",     (0,0), (-1,-1), "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 9),
            ("TEXTCOLOR",    (1,2), (1,2),   sc_g),
            ("TEXTCOLOR",    (3,2), (3,2),   sc_g),
            ("GRID",         (0,0), (-1,-1), 0.5, ICE),
            ("PADDING",      (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS", (0,0), (-1,-1),
             [LIGHT, WHITE, colors.HexColor("#F0FFF8")]),
        ]))
        el.append(rt)
        el.append(Spacer(1, 0.6*cm))

        # Tableau apprenants trié par score
        el.append(Paragraph("Résultats par apprenant", s_h2))
        el.append(HRFlowable(width="100%", thickness=1, color=ICE))
        el.append(Spacer(1, 0.3*cm))

        rows = [["#", "Apprenant", "Taux (%)", "Appréciation"]]
        for i, d in enumerate(sorted(data_app, key=lambda x: -x["percent"]), 1):
            rows.append([str(i), d["nom"], f"{d['percent']} %", score_label(d["percent"])])
        rows.append(["—", "TOTAL SESSION", f"{pct_glob} %", appreci_g])

        at = Table(rows, colWidths=[1*cm, 7*cm, 3*cm, 5.5*cm])
        ts = [
            ("BACKGROUND", (0,0), (-1,0), NAVY),
            ("TEXTCOLOR",  (0,0), (-1,0), WHITE),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,-1), 9),
            ("ALIGN",      (0,0), (-1,-1), "CENTER"),
            ("ALIGN",      (1,1), (1,-1),  "LEFT"),
            ("GRID",       (0,0), (-1,-1), 0.5, ICE),
            ("PADDING",    (0,0), (-1,-1), 7),
            ("BACKGROUND", (0,-1), (-1,-1), BLUE),
            ("TEXTCOLOR",  (0,-1), (-1,-1), WHITE),
            ("FONTNAME",   (0,-1), (-1,-1), "Helvetica-Bold"),
        ]
        for ri2, d in enumerate(sorted(data_app, key=lambda x: -x["percent"]), 1):
            sc = score_color(d["percent"])
            bg = ("#F0FDF4" if d["percent"] >= 75
                  else "#FFFBEB" if d["percent"] >= 50
                  else "#FFF1F2")
            ts += [
                ("TEXTCOLOR",  (2, ri2), (3, ri2), sc),
                ("BACKGROUND", (2, ri2), (3, ri2), colors.HexColor(bg)),
            ]
        at.setStyle(TableStyle(ts))
        el.append(at)

        # ── Graphiques + Analyses ──────────────────────────────
        import base64
        from reportlab.platypus import Image as RLImage

        chart_bars = chart_crit = chart_pie = None
        if request.method == "POST":
            body       = request.data
            chart_bars = body.get("chart_bars")
            chart_crit = body.get("chart_crit")
            chart_pie  = body.get("chart_pie")

        def b64_img(data_url, w_cm, h_cm):
            if not data_url:
                return None
            _, b64 = data_url.split(",", 1)
            return RLImage(io.BytesIO(base64.b64decode(b64)),
                           width=w_cm*cm, height=h_cm*cm)

        s_chart_title = ParagraphStyle(
            "ChartTitle", parent=st["Normal"],
            fontSize=11, fontName="Helvetica-Bold",
            textColor=NAVY, spaceBefore=10, spaceAfter=4,
        )
        s_analysis = ParagraphStyle(
            "AnalysisText", parent=st["Normal"],
            fontSize=9, textColor=colors.HexColor("#4A5A8A"),
            leading=14, spaceBefore=6, spaceAfter=4,
            leftIndent=10, borderPad=4,
        )

        def analysis_text(pct):
            """Génère un commentaire analytique selon le score global."""
            if pct >= 80:
                return (f"✅ <b>Excellente performance</b> — Le taux global de {pct:.1f}% reflète une "
                        "maîtrise solide des compétences évaluées. Les résultats dépassent "
                        "les seuils d'excellence fixés à 75%, ce qui témoigne de l'efficacité "
                        "pédagogique de la formation.")
            elif pct >= 65:
                return (f"✔ <b>Performance satisfaisante</b> — Avec un taux de {pct:.1f}%, "
                        "la formation atteint globalement ses objectifs. Quelques points "
                        "d'amélioration restent identifiables pour les rubriques en deçà du seuil.")
            elif pct >= 50:
                return (f"⚠ <b>Performance passable</b> — Le taux de {pct:.1f}% indique que "
                        "la moitié des apprenants ont atteint le niveau satisfaisant. "
                        "Un renforcement ciblé sur les rubriques faibles est recommandé.")
            else:
                return (f"🔴 <b>Performance insuffisante</b> — Le taux de {pct:.1f}% est en deçà "
                        "des seuils attendus. Une révision du contenu et des méthodes "
                        "pédagogiques est fortement recommandée.")

        def crit_analysis(critere_data):
            """Analyse des rubriques les plus fortes et les plus faibles."""
            if not critere_data:
                return "Aucune donnée de rubrique disponible."
            srt = sorted(critere_data, key=lambda x: x["pourcentage"], reverse=True)
            best  = srt[0]  if len(srt) >= 1 else None
            worst = srt[-1] if len(srt) >= 2 else None
            txt = ""
            if best:
                txt += (f"📊 <b>Rubrique la mieux maîtrisée</b> : « {best['nom']} » avec "
                        f"{best['pourcentage']:.1f}%. ")
            if worst and worst["nom"] != best["nom"]:
                txt += (f"La rubrique nécessitant le plus d'attention est "
                        f"« {worst['nom']} » ({worst['pourcentage']:.1f}%). ")
            below = [c for c in srt if c["pourcentage"] < 50]
            if below:
                noms = ", ".join(f'« {c["nom"]} »' for c in below[:3])
                txt += f"Les rubriques en dessous du seuil minimal (50%) sont : {noms}."
            else:
                txt += "Toutes les rubriques dépassent le seuil minimal de 50%."
            return txt

        def pie_analysis(notes_counter):
            total = sum(notes_counter.values())
            if not total:
                return "Aucune évaluation enregistrée."
            tres = notes_counter.get(3, 0)
            sat  = notes_counter.get(2, 0)
            pas  = notes_counter.get(1, 0)
            p_tres = round(tres/total*100)
            p_sat  = round(sat/total*100)
            p_pas  = round(pas/total*100)
            return (f"📈 <b>{p_tres}% de Très satisfaisant</b>, {p_sat}% de Satisfaisant, "
                    f"{p_pas}% de Pas satisfaisant sur {total} évaluations au total. "
                    + ("Le niveau général est jugé très satisfaisant." if p_tres >= 50
                       else "Une amélioration du niveau général reste possible." if p_sat >= 40
                       else "Des efforts significatifs sont nécessaires pour améliorer le niveau général."))

        # Récupérer les données critères pour l'analyse textuelle
        critere_data_list = []
        for c in criteres:
            evs = session.evaluation_set.filter(critere=c)
            pts = sum(NOTE_MAPPING.get(e.note, 0) for e in evs)
            pct = round(pts / (nb_app * 75) * 100, 1) if nb_app else 0
            critere_data_list.append({"nom": c.nom, "pourcentage": pct})

        notes_counter_dict = {1:0, 2:0, 3:0}
        for ev in session.evaluation_set.all():
            if ev.note in notes_counter_dict:
                notes_counter_dict[ev.note] += 1

        if chart_bars or chart_crit or chart_pie:
            el.append(Spacer(1, 0.8*cm))
            el.append(HRFlowable(width="100%", thickness=2, color=BLUE))
            el.append(Spacer(1, 0.3*cm))
            el.append(Paragraph("Analyses graphiques", s_h2))
            el.append(Spacer(1, 0.3*cm))

        # ── Graphe 1 : Scores apprenants ──
        if chart_bars:
            img = b64_img(chart_bars, 15.5, max(6, nb_app * 0.55 + 2))
            if img:
                el.append(Paragraph("① Scores individuels par apprenant", s_chart_title))
                el.append(HRFlowable(width="100%", thickness=0.5, color=ICE))
                el.append(Spacer(1, 0.2*cm))
                el.append(img)
                el.append(Paragraph(
                    analysis_text(pct_glob),
                    s_analysis
                ))
                el.append(Spacer(1, 0.5*cm))

        # ── Graphe 2 : Scores rubriques ──
        if chart_crit:
            img = b64_img(chart_crit, 15.5, 7.5)
            if img:
                el.append(Paragraph("② Performance par rubrique d'évaluation", s_chart_title))
                el.append(HRFlowable(width="100%", thickness=0.5, color=ICE))
                el.append(Spacer(1, 0.2*cm))
                el.append(img)
                el.append(Paragraph(
                    crit_analysis(critere_data_list),
                    s_analysis
                ))
                el.append(Spacer(1, 0.5*cm))

        # ── Graphe 3 : Satisfaction ──
        if chart_pie:
            img = b64_img(chart_pie, 14, 6.5)
            if img:
                el.append(Paragraph("③ Répartition des niveaux de satisfaction", s_chart_title))
                el.append(HRFlowable(width="100%", thickness=0.5, color=ICE))
                el.append(Spacer(1, 0.2*cm))
                el.append(img)
                el.append(Paragraph(
                    pie_analysis(notes_counter_dict),
                    s_analysis
                ))

        # ── Commentaires individuels (depuis le frontend — filtrés par l'utilisateur) ──
        commentaires_app = []
        if request.method == "POST":
            for item in request.data.get("commentaires_apprenants", []):
                nom = item.get("nom","").strip()
                txt = item.get("commentaire","").strip()
                if nom and txt:
                    commentaires_app.append((nom, txt))

        if commentaires_app:
            el.append(Spacer(1, 0.8*cm))
            el.append(HRFlowable(width="100%", thickness=2, color=BLUE))
            el.append(Spacer(1, 0.3*cm))
            el.append(Paragraph("Commentaires des apprenants", s_h2))
            el.append(Spacer(1, 0.2*cm))
            s_comment_name = ParagraphStyle(
                "CommentName", parent=st["Normal"],
                fontSize=10, fontName="Helvetica-Bold",
                textColor=NAVY, spaceBefore=6, spaceAfter=2,
            )
            s_comment_body = ParagraphStyle(
                "CommentBody", parent=st["Normal"],
                fontSize=9, textColor=colors.HexColor("#4A5A8A"),
                leading=14, leftIndent=12, spaceAfter=6,
                borderPad=4,
            )
            for nom, comment in commentaires_app:
                el.append(Paragraph(f"▸ {nom}", s_comment_name))
                el.append(Paragraph(comment, s_comment_body))
                el.append(HRFlowable(width="90%", thickness=0.5, color=ICE))

        # ── Commentaire final de session ──────────────────────
        commentaire_final = ""
        if request.method == "POST":
            commentaire_final = request.data.get("commentaire_final", "").strip()
        if not commentaire_final:
            try:
                commentaire_final = session.commentaire_final or ""
            except Exception:
                commentaire_final = ""
        if commentaire_final:
            el.append(Spacer(1, 0.8*cm))
            el.append(HRFlowable(width="100%", thickness=2, color=BLUE))
            el.append(Spacer(1, 0.3*cm))
            el.append(Paragraph("Commentaire final", s_h2))
            el.append(Spacer(1, 0.2*cm))
            s_final = ParagraphStyle(
                "CommentFinal", parent=st["Normal"],
                fontSize=10, textColor=colors.HexColor("#0D1B5E"),
                leading=16, leftIndent=10, spaceAfter=8,
                backColor=colors.HexColor("#F0F4FF"),
                borderPad=10, borderRadius=4,
            )
            el.append(Paragraph(commentaire_final, s_final))

        doc.build(el)
        buf.seek(0)
        return HttpResponse(buf, content_type="application/pdf")

















































#Inscription candidat View
# views.py  (section Candidat)
# views.py
# views.py
# views.py
import logging
from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q, Count, Avg

from .models import Candidat, Formation
from .serializers import (
    CandidatSerializer,
    CandidatLightSerializer,
    FormationSerializer,
)

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════
#  VIEWSET FORMATION
# ══════════════════════════════════════════════════════════════════

class FormationViewSet(viewsets.ModelViewSet):
    """
    CRUD complet pour les formations ONFPP.

    GET    /api/formations/                → liste (+ filtres)
    POST   /api/formations/                → créer
    GET    /api/formations/{id}/           → détail
    PATCH  /api/formations/{id}/           → modifier
    DELETE /api/formations/{id}/           → supprimer
    GET    /api/formations/{id}/candidats/ → apprenants de la session
    GET    /api/formations/stats/          → statistiques
    """

    queryset           = Formation.objects.all().order_by("-created_at")
    serializer_class   = FormationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = Formation.objects.annotate(
            _nb_candidats=Count("candidats")
        ).order_by("-created_at")

        search = self.request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(nom_formation__icontains=search) |
                Q(organisme_formation__icontains=search) |
                Q(nom_formateur__icontains=search)
            )

        antenne = self.request.query_params.get("antenne", "").strip()
        if antenne:
            qs = qs.filter(antenne=antenne)

        type_f = self.request.query_params.get("type", "").strip()
        if type_f in ("continue", "apprentissage"):
            qs = qs.filter(type_formation=type_f)

        division = self.request.query_params.get("division", "").strip().upper()
        if division == "DAP":
            qs = qs.filter(type_formation="apprentissage")
        elif division == "DFC":
            qs = qs.filter(type_formation="continue")

        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["get"], url_path="candidats")
    def candidats(self, request, pk=None):
        """
        GET /api/formations/{id}/candidats/
        Retourne les apprenants de cette session avec stats insertion.
        """
        formation = self.get_object()
        qs = formation.candidats.all().order_by("-created_at")

        statut = request.query_params.get("statut", "").strip()
        if statut in ("en_attente", "valide", "rejete"):
            qs = qs.filter(statut_fiche=statut)

        insere = request.query_params.get("insere", "").strip()
        if insere == "true":
            qs = qs.filter(insere=True)
        elif insere == "false":
            qs = qs.filter(Q(insere=False) | Q(insere__isnull=True))

        serializer = CandidatLightSerializer(qs, many=True)
        return Response({
            "formation_id":          formation.id,
            "nom_formation":         formation.nom_formation,
            "identifiant_formation": formation.identifiant_formation,
            "division":              formation.division,
            "antenne":               formation.get_antenne_display(),
            "nb_candidats":          qs.count(),
            "nb_inseres":            qs.filter(insere=True).count(),
            "candidats":             serializer.data,
        })

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        qs = Formation.objects.all()
        return Response({
            "total":            qs.count(),
            "dap":              qs.filter(type_formation="apprentissage").count(),
            "dfc":              qs.filter(type_formation="continue").count(),
            "total_apprenants": Candidat.objects.filter(formation__isnull=False).count(),
            "par_antenne": list(
                qs.values("antenne").annotate(nb=Count("id")).order_by("-nb")
            ),
            "par_division": [
                {"division": "DAP", "nb": qs.filter(type_formation="apprentissage").count()},
                {"division": "DFC", "nb": qs.filter(type_formation="continue").count()},
            ],
        })


# ══════════════════════════════════════════════════════════════════
#  VIEWSET CANDIDAT / APPRENANT
# ══════════════════════════════════════════════════════════════════

class CandidatViewSet(viewsets.ModelViewSet):
    """
    CRUD complet + actions métier.

    GET    /api/candidats/                    → liste (filtres avancés)
    POST   /api/candidats/                    → créer (+ formation_data imbriqué)
    GET    /api/candidats/{id}/               → détail
    PATCH  /api/candidats/{id}/               → modifier
    DELETE /api/candidats/{id}/               → supprimer
    PATCH  /api/candidats/{id}/valider/       → valider fiche + génère ID
    PATCH  /api/candidats/{id}/rejeter/       → rejeter fiche
    PATCH  /api/candidats/{id}/inserer/       → marquer comme inséré (+ détails)
    PATCH  /api/candidats/{id}/desinserer/    → annuler insertion
    GET    /api/candidats/stats/              → statistiques globales
    GET    /api/candidats/stats_insertion/    → stats insertion détaillées

    Filtres disponibles :
      ?search=     → nom, prénom, email, téléphone, identifiant
      ?antenne=    → code antenne
      ?formation=  → id formation
      ?domaine=    → code domaine
      ?situation=  → code situation
      ?sexe=       → H | F
      ?statut=     → en_attente | valide | rejete
      ?insere=     → true | false
      ?statut_emploi= → en_poste | chomage | independant | etudes | autre
    """

    queryset = (
        Candidat.objects
        .select_related("formation", "created_by")
        .order_by("-created_at")
    )
    serializer_class   = CandidatSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = (
            Candidat.objects
            .select_related("formation", "created_by")
            .order_by("-created_at")
        )

        search = self.request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(nom__icontains=search) |
                Q(prenom__icontains=search) |
                Q(email__icontains=search) |
                Q(telephone__icontains=search) |
                Q(identifiant_unique__icontains=search) |
                Q(formation__nom_formation__icontains=search) |
                Q(entreprise_insertion__icontains=search) |
                Q(poste_occupe__icontains=search)
            )

        antenne = self.request.query_params.get("antenne", "").strip()
        if antenne:
            qs = qs.filter(antenne=antenne)

        formation_id = self.request.query_params.get("formation", "").strip()
        if formation_id:
            qs = qs.filter(formation_id=formation_id)

        domaine = self.request.query_params.get("domaine", "").strip()
        if domaine:
            qs = qs.filter(domaine=domaine)

        situation = self.request.query_params.get("situation", "").strip()
        if situation:
            qs = qs.filter(situation=situation)

        sexe = self.request.query_params.get("sexe", "").strip()
        if sexe in ("H", "F"):
            qs = qs.filter(sexe=sexe)

        statut = self.request.query_params.get("statut", "").strip()
        if statut in ("en_attente", "valide", "rejete"):
            qs = qs.filter(statut_fiche=statut)

        insere = self.request.query_params.get("insere", "").strip()
        if insere == "true":
            qs = qs.filter(insere=True)
        elif insere == "false":
            qs = qs.filter(Q(insere=False) | Q(insere__isnull=True))

        statut_emploi = self.request.query_params.get("statut_emploi", "").strip()
        if statut_emploi:
            qs = qs.filter(statut_emploi_actuel=statut_emploi)

        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    # ── Valider une fiche ─────────────────────────────────────────

    @action(detail=True, methods=["patch"], url_path="valider")
    def valider(self, request, pk=None):
        """PATCH /api/candidats/{id}/valider/ — Valide la fiche et génère l'identifiant."""
        candidat = self.get_object()
        if candidat.statut_fiche == "valide":
            return Response({"detail": "Ce candidat est déjà validé."},
                            status=status.HTTP_400_BAD_REQUEST)
        candidat.statut_fiche = "valide"
        if not candidat.identifiant_unique:
            candidat.identifiant_unique = candidat.generate_identifiant()
        candidat.save()
        return Response(CandidatSerializer(candidat, context={"request": request}).data)

    # ── Rejeter une fiche ─────────────────────────────────────────

    @action(detail=True, methods=["patch"], url_path="rejeter")
    def rejeter(self, request, pk=None):
        """PATCH /api/candidats/{id}/rejeter/"""
        candidat = self.get_object()
        if candidat.statut_fiche == "rejete":
            return Response({"detail": "Ce candidat est déjà rejeté."},
                            status=status.HTTP_400_BAD_REQUEST)
        candidat.statut_fiche = "rejete"
        candidat.save()
        return Response(CandidatSerializer(candidat, context={"request": request}).data)

    # ── Marquer comme inséré ──────────────────────────────────────

    @action(detail=True, methods=["patch"], url_path="inserer")
    def inserer(self, request, pk=None):
        """
        PATCH /api/candidats/{id}/inserer/

        Body optionnel (tous les champs insertion) :
        {
          "entreprise_insertion": "Société ABC",
          "poste_occupe": "Technicien réseau",
          "type_contrat": "cdi",
          "salaire_insertion": 2500000,
          "secteur_activite": "Télécommunications",
          "date_insertion": "2026-02-01",
          "duree_recherche_emploi": 3,
          "statut_emploi_actuel": "en_poste",
          "commentaire_insertion": "Très bonne intégration",
          "date_suivi_insertion": "2026-03-01",
          "suivi_par": "Conseiller Diallo"
        }
        """
        candidat = self.get_object()
        insertion_fields = [
            "entreprise_insertion", "poste_occupe", "type_contrat",
            "salaire_insertion", "secteur_activite", "date_insertion",
            "duree_recherche_emploi", "formation_complementaire",
            "satisfaction_formation", "statut_emploi_actuel",
            "commentaire_insertion", "date_suivi_insertion", "suivi_par",
        ]
        candidat.insere = True
        for field in insertion_fields:
            if field in request.data:
                setattr(candidat, field, request.data[field])
        # Défaut statut emploi
        if not candidat.statut_emploi_actuel:
            candidat.statut_emploi_actuel = "en_poste"
        candidat.save()
        return Response(CandidatSerializer(candidat, context={"request": request}).data)

    # ── Annuler l'insertion ───────────────────────────────────────

    @action(detail=True, methods=["patch"], url_path="desinserer")
    def desinserer(self, request, pk=None):
        """PATCH /api/candidats/{id}/desinserer/ — Remet insere=False."""
        candidat = self.get_object()
        candidat.insere = False
        candidat.statut_emploi_actuel = request.data.get("statut_emploi_actuel", "chomage")
        candidat.commentaire_insertion = request.data.get("commentaire_insertion",
                                                           candidat.commentaire_insertion)
        candidat.save()
        return Response(CandidatSerializer(candidat, context={"request": request}).data)

    # ── Statistiques globales ─────────────────────────────────────

    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        qs = Candidat.objects.all()
        return Response({
            "total":      qs.count(),
            "en_attente": qs.filter(statut_fiche="en_attente").count(),
            "valide":     qs.filter(statut_fiche="valide").count(),
            "rejete":     qs.filter(statut_fiche="rejete").count(),
            "inseres":    qs.filter(insere=True).count(),
            "par_situation": list(
                qs.exclude(situation__isnull=True)
                .values("situation").annotate(nb=Count("id")).order_by("-nb")
            ),
            "par_domaine": list(
                qs.exclude(domaine__isnull=True)
                .values("domaine").annotate(nb=Count("id")).order_by("-nb")
            ),
            "par_antenne": list(
                qs.exclude(antenne__isnull=True)
                .values("antenne").annotate(nb=Count("id")).order_by("-nb")
            ),
            "par_sexe": list(
                qs.exclude(sexe__isnull=True)
                .values("sexe").annotate(nb=Count("id"))
            ),
            "par_niveau_etude": list(
                qs.exclude(niveau_etude__isnull=True)
                .values("niveau_etude").annotate(nb=Count("id")).order_by("-nb")
            ),
        })

    # ── Statistiques insertion détaillées ─────────────────────────

    @action(detail=False, methods=["get"], url_path="stats_insertion")
    def stats_insertion(self, request):
        """
        GET /api/candidats/stats_insertion/
        Tableau de bord complet de l'insertion professionnelle.
        """
        qs = Candidat.objects.filter(statut_fiche="valide")
        total_valides = qs.count()
        inseres_qs    = qs.filter(insere=True)
        nb_inseres    = inseres_qs.count()

        taux = round((nb_inseres / total_valides * 100), 1) if total_valides else 0

        return Response({
            "total_valides":     total_valides,
            "nb_inseres":        nb_inseres,
            "nb_non_inseres":    total_valides - nb_inseres,
            "taux_insertion":    taux,

            "par_type_contrat": list(
                inseres_qs.exclude(type_contrat__isnull=True)
                .values("type_contrat").annotate(nb=Count("id")).order_by("-nb")
            ),
            "par_statut_emploi": list(
                qs.exclude(statut_emploi_actuel__isnull=True)
                .values("statut_emploi_actuel").annotate(nb=Count("id")).order_by("-nb")
            ),
            "par_domaine": list(
                inseres_qs.exclude(domaine__isnull=True)
                .values("domaine").annotate(nb=Count("id")).order_by("-nb")
            ),
            "par_antenne": list(
                inseres_qs.exclude(antenne__isnull=True)
                .values("antenne").annotate(nb=Count("id")).order_by("-nb")
            ),
            "par_secteur": list(
                inseres_qs.exclude(secteur_activite__isnull=True)
                .values("secteur_activite").annotate(nb=Count("id")).order_by("-nb")[:10]
            ),
            "salaire_moyen":              inseres_qs.exclude(salaire_insertion__isnull=True)
                                          .aggregate(avg=Avg("salaire_insertion"))["avg"],
            "duree_recherche_moyenne":    inseres_qs.exclude(duree_recherche_emploi__isnull=True)
                                          .aggregate(avg=Avg("duree_recherche_emploi"))["avg"],
            "satisfaction_moyenne":       qs.exclude(satisfaction_formation__isnull=True)
                                          .aggregate(avg=Avg("satisfaction_formation"))["avg"],
            "avec_formation_complementaire": qs.filter(formation_complementaire=True).count(),
        })



































# À ajouter dans views.py

from rest_framework import viewsets, permissions, status
from rest_framework.decorators import action
from rest_framework.response import Response
from django.db.models import Q, Count, Sum, Avg

from .models import EntrepriseFormation, ModulePlanFormation
from .serializers import (
    EntrepriseFormationSerializer,
    EntrepriseFormationLightSerializer,
    ModulePlanFormationSerializer,
)


class EntrepriseFormationViewSet(viewsets.ModelViewSet):
    """
    CRUD complet pour les plans de formation entreprises (DFC).

    GET    /api/entreprise-formations/           → liste
    POST   /api/entreprise-formations/           → créer
    GET    /api/entreprise-formations/{id}/      → détail (avec modules)
    PATCH  /api/entreprise-formations/{id}/      → modifier
    DELETE /api/entreprise-formations/{id}/      → supprimer

    Actions métier :
    GET  /api/entreprise-formations/{id}/modules/           → modules du plan
    POST /api/entreprise-formations/{id}/add_module/        → ajouter un module
    PATCH /api/entreprise-formations/{id}/update_statut/    → changer statut
    GET  /api/entreprise-formations/stats/                  → statistiques globales

    Filtres :
      ?search=        → nom_entreprise, intitule_formation, identifiant_unique
      ?antenne=       → code antenne
      ?statut=        → planifiee | en_cours | realisee | annulee | reportee
      ?annee=         → année de soumission (ex: 2026)
    """

    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action == "list":
            return EntrepriseFormationLightSerializer
        return EntrepriseFormationSerializer

    def get_queryset(self):
        qs = EntrepriseFormation.objects.select_related(
            "created_by", "session_evaluation"
        ).prefetch_related("modules").order_by("-created_at")

        search = self.request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(nom_entreprise__icontains=search) |
                Q(intitule_formation__icontains=search) |
                Q(identifiant_unique__icontains=search) |
                Q(secteur_activite__icontains=search) |
                Q(contact_rh__icontains=search)
            )

        antenne = self.request.query_params.get("antenne", "").strip()
        if antenne:
            qs = qs.filter(antenne=antenne)

        statut = self.request.query_params.get("statut", "").strip()
        if statut:
            qs = qs.filter(statut_realisation=statut)

        annee = self.request.query_params.get("annee", "").strip()
        if annee.isdigit():
            qs = qs.filter(date_soumission__year=int(annee))

        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    # ── GET modules d'un plan ─────────────────────────────────────
    @action(detail=True, methods=["get"], url_path="modules")
    def modules(self, request, pk=None):
        entreprise_formation = self.get_object()
        modules = entreprise_formation.modules.all().order_by("ordre", "date_debut_prevue")
        serializer = ModulePlanFormationSerializer(modules, many=True)
        return Response({
            "identifiant":   entreprise_formation.identifiant_unique,
            "entreprise":    entreprise_formation.nom_entreprise,
            "intitule":      entreprise_formation.intitule_formation,
            "nb_modules":    modules.count(),
            "nb_planifies":  modules.filter(statut="planifiee").count(),
            "nb_en_cours":   modules.filter(statut="en_cours").count(),
            "nb_realises":   modules.filter(statut="realisee").count(),
            "modules":       serializer.data,
        })

    # ── POST ajouter un module ────────────────────────────────────
    @action(detail=True, methods=["post"], url_path="add-module")
    def add_module(self, request, pk=None):
        entreprise_formation = self.get_object()
        serializer = ModulePlanFormationSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(entreprise_formation=entreprise_formation)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # ── PATCH mettre à jour statut ────────────────────────────────
    @action(detail=True, methods=["patch"], url_path="update-statut")
    def update_statut(self, request, pk=None):
        obj = self.get_object()
        new_statut = request.data.get("statut_realisation")
        valid = [c[0] for c in EntrepriseFormation.STATUT_REALISATION_CHOICES
                 if hasattr(EntrepriseFormation, 'STATUT_REALISATION_CHOICES')]

        # Utiliser les choix directement du modèle
        valid_statuts = ["planifiee", "en_cours", "realisee", "annulee", "reportee"]
        if new_statut not in valid_statuts:
            return Response(
                {"detail": f"Statut invalide. Valeurs possibles : {valid_statuts}"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        obj.statut_realisation = new_statut
        if new_statut == "realisee" and request.data.get("date_realisation"):
            obj.date_realisation = request.data["date_realisation"]
        obj.save()
        return Response(EntrepriseFormationSerializer(obj, context={"request": request}).data)

    # ── GET stats globales ────────────────────────────────────────
    @action(detail=False, methods=["get"], url_path="stats")
    def stats(self, request):
        qs = EntrepriseFormation.objects.all()
        inseres = qs.aggregate(
            total_hommes=Sum("nb_hommes"),
            total_femmes=Sum("nb_femmes"),
            total_formes_h=Sum("nb_formes_hommes"),
            total_formes_f=Sum("nb_formes_femmes"),
        )
        return Response({
            "total":             qs.count(),
            "planifiees":        qs.filter(statut_realisation="planifiee").count(),
            "en_cours":          qs.filter(statut_realisation="en_cours").count(),
            "realisees":         qs.filter(statut_realisation="realisee").count(),
            "annulees":          qs.filter(statut_realisation="annulee").count(),
            "reportees":         qs.filter(statut_realisation="reportee").count(),
            "total_employes_prevus": (inseres["total_hommes"] or 0) + (inseres["total_femmes"] or 0),
            "total_hommes_prevus":   inseres["total_hommes"] or 0,
            "total_femmes_prevus":   inseres["total_femmes"] or 0,
            "total_formes":          (inseres["total_formes_h"] or 0) + (inseres["total_formes_f"] or 0),
            "total_formes_hommes":   inseres["total_formes_h"] or 0,
            "total_formes_femmes":   inseres["total_formes_f"] or 0,
            "par_antenne": list(
                qs.values("antenne").annotate(nb=Count("id")).order_by("-nb")
            ),
            "par_statut": list(
                qs.values("statut_realisation").annotate(nb=Count("id")).order_by("-nb")
            ),
            "par_secteur": list(
                qs.exclude(secteur_activite__isnull=True)
                .values("secteur_activite").annotate(nb=Count("id")).order_by("-nb")[:8]
            ),
        })


# ══════════════════════════════════════════════════════════════════
#  VIEWSET MODULE
# ══════════════════════════════════════════════════════════════════

class ModulePlanFormationViewSet(viewsets.ModelViewSet):
    """
    CRUD sur les modules d'un plan de formation.
    GET/PATCH/DELETE /api/modules-formation/{id}/
    """
    queryset           = ModulePlanFormation.objects.all()
    serializer_class   = ModulePlanFormationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = super().get_queryset()
        plan_id = self.request.query_params.get("plan")
        if plan_id:
            qs = qs.filter(entreprise_formation_id=plan_id)
        return qs

    @action(detail=True, methods=["patch"], url_path="update-statut")
    def update_statut(self, request, pk=None):
        module = self.get_object()
        new_statut = request.data.get("statut")
        valid = ["planifiee", "en_cours", "realisee", "annulee", "reportee"]
        if new_statut not in valid:
            return Response({"detail": f"Statut invalide."}, status=400)
        module.statut = new_statut
        if new_statut == "realisee" and request.data.get("date_realisation"):
            module.date_realisation = request.data["date_realisation"]
        module.save()
        return Response(ModulePlanFormationSerializer(module).data)